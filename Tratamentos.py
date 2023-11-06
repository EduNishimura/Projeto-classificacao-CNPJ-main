import openpyxl
import pandas as pd


def tirando_separadores_e_none(arquivo, plani):
    """
            Remove linhas com valores vazios (None) e divide células contendo vírgulas em várias linhas.

            Parâmetros:
                arquivo (str): Nome do arquivo Excel.
                plani (str): Nome da planilha a ser processada.
    """
    arquivos = openpyxl.load_workbook(arquivo)
    planilha = arquivos[plani]
    for c in range(len(planilha["A"]), 1, -1):
        cedula = planilha[f'A{c}'].value
        if cedula is None:
            planilha.delete_rows(c)
        elif ',' in str(cedula):
            partes = cedula.split(",")
            for parte in range(0, len(partes)):
                planilha.insert_rows(c + 1)
                planilha[f'A{c + 1}'].value = partes[parte]
            planilha.delete_rows(c)
            arquivos.save(arquivo)
    print("Separadores Feitos")


def tirando_duplicatas_e_espacos(arquivo, plani):
    """
        Remove valores duplicados e espaços no início ou no fim da coluna especificada na planilha do Excel.

        Parâmetros:
            arquivo (str): Nome do arquivo Excel.
            plani (str): Nome da planilha a ser processada.
    """
    df = pd.read_excel(arquivo, sheet_name=plani)
    df.iloc[:, 0] = df.iloc[:, 0].str.strip()
    df_sem_duplicatas = df.drop_duplicates(keep='first')
    df_sem_duplicatas.to_excel(arquivo, index=False, sheet_name=plani)
    print("Duplicatas retiradas")


def tirando_cei(arquivo, plani):
    """
       Remove linhas cujo conteúdo da primeira coluna tenha um comprimento inferior a 16 caracteres.

       Parâmetros:
           arquivo (str): Nome do arquivo Excel.
           plani (str): Nome da planilha a ser processada.
   """
    df = pd.read_excel(arquivo, sheet_name=plani)
    df = df[df.iloc[:, 0].str.len() >= 16]
    df.to_excel(arquivo, sheet_name=plani, index=False)
    print("Cei retirados")


def adicionando_cnae(arquivo):
    """
        Adiciona uma nova planilha 'CNAE' ao arquivo Excel e preenche-a com dados do arquivo 'CNAE.xlsx'.

        Parâmetros:
            arquivo (str): Nome do arquivo Excel.
    """
    arquivo_principal = openpyxl.load_workbook(arquivo)
    arquivo_cnae = openpyxl.load_workbook('CNAE.xlsx')
    planilha_cnae = arquivo_cnae['CNAE']
    copy = arquivo_principal.create_sheet(title='CNAE')
    for row in planilha_cnae.iter_rows(values_only=True):
        copy.append(row)
    arquivo_principal.save(arquivo)
    print("planilha CNAE Adicionada")


'''
Exemplo de usos:
tirando_separadores_e_none(arquivo='CNPJ - Trabalhador.xlsx', plani='Planilha1')
tirando_duplicatas_e_espacos(arquivo='CNPJ - Trabalhador.xlsx', plani='Planilha1')
tirando_cei(arquivo='CNPJ - Trabalhador.xlsx', plani='Planilha1')
adicionando_cnae(arquivo='CNPJ - Trabalhador.xlsx')

Importante, para rodar o mais eles devem estar em modo de leitura (por isso os 3 aspas)
# <-- observações de uma linha
aspas triplo <-- Observações de várias linhas, como esse.
'''
